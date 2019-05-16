#!python3
import urllib.request
import re
import time
import string
import datetime
import openpyxl
import os

'''
TODO arbeidsinkomen berekening niet hardcoden
TODO heffinsvrijvermogen https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/vermogen_en_aanmerkelijk_belang/vermogen/belasting_betalen_over_uw_vermogen/heffingsvrij_vermogen/
TODO premievolk
TODO roken/alcohol accijns
'''


class Persoon:

    def __init__(self, postcode, leeftijd, bruto_loon_mnd, bonus, uitgaven_laag, uitgaven_hoog, spaargeld, schulden,
                 verbruik_gas, verbruik_stroom, verbruik_water):
        params_init = locals()
        params = check_input(params_init, [0], range(1, 11))

        self.postcode = params['postcode']
        self.leeftijd = params['leeftijd']
        self.bruto_loon_mnd = params['bruto_loon_mnd']
        self.bonus = params['bonus']
        self.uitgaven_laag = params['uitgaven_laag'] * 12
        self.uitgaven_hoog = params['uitgaven_hoog'] * 12
        self.spaargeld = params['spaargeld']
        self.schulden = params['schulden']
        self.verbruik_gas = params['verbruik_gas']
        self.verbruik_stroom = params['verbruik_stroom']
        self.verbruik_water = params['verbruik_water']
        self.bruto_loon_jr = self.bruto_loon_mnd * 12
        self.vakantiegeld = self.bruto_loon_jr * 0.08
        self.roken = 0
        self.alcohol = 0
        self.loon_speciaal = self.vakantiegeld + self.bonus
        self.loon_totaal = self.bruto_loon_jr + self.loon_speciaal

        print('Getting location data ...')
        start = time.perf_counter()
        url = 'https://www.postcodezoekmachine.nl/' + self.postcode.upper()
        response = urllib.request.urlopen(url)
        html = response.read().decode(response.headers.get_content_charset())

        self.stad = regex_lookup("<td><a href=\"\/(\w*)\">", html)
        self.stad = string.capwords(self.stad, '-')
        self.provincie = regex_lookup("<td><a href=\"\/provincie\/(.*?)\">", html)
        self.provincie = string.capwords(self.provincie, '-')
        self.gemeente = regex_lookup("<td><a href=\"\/gemeente\/(.*?)\">", html)
        self.gemeente = string.capwords(self.gemeente, '-')
        print('Success! Data found in {} seconds'.format(round(time.perf_counter() - start, 2)))
        print('{}, {}, {}'.format(self.stad, self.provincie, self.gemeente))

    def roken(self, roken_per_week):
        self.roken = roken_per_week

    def drinken(self, alcohol_per_week):
        self.alcohol = alcohol_per_week


class Voertuig:

    def __init__(self, persoon, kenteken, prijs, km_jaar):
        params_init = locals()
        params = check_input(params_init, [1], [2, 3])

        self.persoon = persoon
        self.kenteken = params['kenteken']
        self.prijs = params['prijs']
        self.km_jaar = params['km_jaar']

        if not (len(self.kenteken) != 8 or len(self.kenteken) != 6):
            self.kenteken = input('Kenteken moet 6 (12abcd) of 8 (12-ab-cd) tekens zijn >>> ')

    def get_data(self):
        # Convert kenteken, length is either 6 or 8
        if len(self.kenteken) == 8:
            self.kenteken = self.kenteken.upper()
        else:  # is 6
            int_count = 0
            for letter in self.kenteken:
                try:
                    int(letter)
                    int_count += 1
                except ValueError:
                    pass
            if int_count > 2:
                self.kenteken = (self.kenteken[:2] + '-' + self.kenteken[2:5] + '-' + self.kenteken[5]).upper()
            else:
                self.kenteken = (self.kenteken[:2] + '-' + self.kenteken[2:4] + '-' + self.kenteken[4:6]).upper()

        print('Getting vehicle data ...')
        start = time.perf_counter()

        # Get html data
        ini_url = 'https://www.kentekencheck.nl/kenteken?i=' + self.kenteken
        ini_html = read_html(ini_url)
        time.sleep(1)
        link_to_rapport = regex_lookup("<iframe src=\"(.*?)\"", ini_html)
        html = read_html('https://www.kentekencheck.nl' + link_to_rapport)

        # Read stuff
        self.merk_nummer = regex_lookup("Merk<\/td>\s*<td style=\"width:60%;\">(.*?)<\/td>", html) + ' ' + regex_lookup(
            "<td>Type<\/td>\s*<td>(.*?)<\/td>", html)
        self.bouwjaar = regex_lookup("<td>Bouwjaar<\/td>\s*<td>.*?(\d{4})<\/td>", html)
        self.brandstof = regex_lookup("<td>Brandstof<\/td>\s*<td>(.*?)<\/td>", html)
        self.nieuwprijs = regex_lookup("<td>Nieuwprijs<\/td>\s*<td>&euro; (.*?)<\/td>", html).replace('.', '')
        self.gewicht = regex_lookup("<td>Massa ledig voertuig<\/td>\s*<td>(\d*) KG<\/td>", html)
        self.wegenbelasting = int(
            regex_lookup("<td>" + self.persoon.provincie + "<\/td>\s*.*?<\/td>\s*<td>&euro;(\d*)<\/td>", html))
        try:
            self.verbruik = round(
                1 / float(regex_lookup("<td>Verbruik gecombineerd<\/td>\s*<td>.*\(1:(.*?)km\)<\/td>", html)), 4)
        except AttributeError:
            print('Verbruik niet gevonden')
        try:
            self.CO2_uitstoot = int(regex_lookup("<td>CO2 uitstoot<\/td>\s*<td>(\d*?) g\/km<\/td>", html))
        except AttributeError:
            print('CO2 uitstoot niet gevonden')

        print('Success! Data found in {} seconds'.format(round(time.perf_counter() - start, 2)))
        print('{}, {}, {}, {}, {}, {}, {}, {}'.format(self.merk_nummer, self.bouwjaar, self.brandstof, self.nieuwprijs,
                                                      self.gewicht, self.wegenbelasting, str(self.verbruik),
                                                      str(self.CO2_uitstoot)))


class Belasting:

    def __init__(self, persoon, auto, huishouden_personen, fiscale_partner):
        self.persoon = persoon
        self.auto = auto
        self.fiscale_partner = fiscale_partner
        try:
            self.huishouden_personen = int(huishouden_personen)
        except ValueError:
            self.huishouden_personen = input('Vul een geldig getal in voor het aantal personen in huis: >>> ')

        current_year = str(datetime.datetime.now().year)

        # Auto
        print('Getting car...')
        url = "https://www.unitedconsumers.com/tanken/informatie/opbouw-brandstofprijzen.asp"
        html = read_html(url)
        self.benzine_prijs = float(regex_lookup("Opbouw Benzine .*?<strong>(.*?)<\/strong>", html).replace(',', '.'))
        self.benzine_btw = int(regex_lookup("Opbouw Benzine .*?BTW.*?(\d*)%", html)) / 100
        self.benzine_accijns = int(regex_lookup("Opbouw Benzine .*?Accijns.*?(\d*)%", html)) / 100
        self.diesel_prijs = float(regex_lookup("Opbouw Diesel .*?<strong>(.*?)<\/strong>", html).replace(',', '.'))
        self.diesel_btw = int(regex_lookup("Opbouw Diesel .*?BTW.*?(\d*)%", html)) / 100
        self.diesel_accijns = int(regex_lookup("Opbouw Diesel .*?Accijns.*?(\d*)%", html)) / 100
        self.MRB = self.auto.wegenbelasting

        # Gemeente
        print('Getting gemeente...')
        url = "https://www.coelo.nl/images/Gemeentelijke_belastingen_{}.xlsx".format(current_year)
        url_name = url.split('/')[-1]
        url_opcenten = "https://www.coelo.nl/images/Provinciale_opcenten_{}.xlsx".format(current_year)
        url_opcenten_name = url_opcenten.split('/')[-1]
        if not os.path.abspath(url_name):
            urllib.request.urlretrieve(url, url_name)
        if not os.path.abspath(url_opcenten_name):
            urllib.request.urlretrieve(url_opcenten, url_opcenten_name)

        wb = openpyxl.load_workbook(os.path.abspath(url_name))['Gegevens per gemeente']
        if self.huishouden_personen > 1:
            col = ['I', 'N', 'S']
        else:
            col = ['I', 'L', 'Q']
        for row in range(5, wb.max_row + 1):
            if wb['B' + str(row)].value == self.persoon.gemeente:
                self.OZB = round(wb[col[0] + str(row)].value, 2)
                self.afvalheffing = round(wb[col[1] + str(row)].value, 2)
                self.rioolheffing = wb[col[2] + str(row)].value
                break

        # Opcenten
        wb = openpyxl.load_workbook(os.path.abspath(url_opcenten_name))['Gegevens per provincie']
        for row in range(6, 18):
            if wb['B' + str(row)].value == self.persoon.provincie:
                self.opcenten = round(wb['C' + str(row)].value / 100, 4)
                break

        # Loonschaal
        url = "https://www.belastingdienst.nl/bibliotheek/handboeken/html/boeken/HL/handboek_loonheffingen_2019-tarieven_bedragen_en_percentages.html"
        html = read_html(url)

        self.loonbelasting_schaal = [
            0,
            int(regex_lookup("rmkrnpakgd.*?t\/m € (.*?)<\/p>", html).replace('.', '')),
            int(regex_lookup("bdoeboonge.*?t\/m € (.*?)<\/p>", html).replace('.', '')),
            int(regex_lookup("eablhjemgh.*?t\/m € (.*?)<\/p>", html).replace('.', ''))
        ]
        self.loonbelasting = [
            float(regex_lookup("obcfqdbaga\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("ehdmneflgk\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("eqqaokdfgo\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("eoadoaopgf\">(.*?)%", html).replace(',', '.')) / 100,
        ]
        self.loonbelasting_aow = [
            float(regex_lookup("pdncrhorgl\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("bkmcoadagj\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("ldfkdkfqge\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("khpffjqjgj\">(.*?)%", html).replace(',', '.')) / 100
        ]

        # Vermogensbelasting
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/vermogen_en_aanmerkelijk_belang/vermogen/belasting_betalen_over_uw_vermogen/grondslag_sparen_en_beleggen/berekening-2019/berekening-belasting-over-inkomen-uit-vermogen-over-2019"
        html = read_html(url)

        self.vermogensbelasting_schaal = [
            0,
            int(regex_lookup("Tot en met €&nbsp;(.*?)<\/p>", html).replace('.', '')),
            int(regex_lookup("Vanaf €&nbsp;.*?tot en met €&nbsp;(.*?)<\/p>", html).replace('.', ''))
        ]
        # self.vermogensbelasting_percentage = [
        #     int(regex_lookup_nogroup("<p>(\d*?)%<\/p>", html)[0]) / 100,
        #     int(regex_lookup_nogroup("<p>(\d*?)%<\/p>", html)[2]) / 100,
        #     0
        # ]
        self.vermogensbelasting_rendement = [
            round(float(regex_lookup_nogroup("<p>(\d*,\d*)%<\/p>", html)[0].replace(',', '.')) / 100, 5),
            round(float(regex_lookup_nogroup("<p>(\d*,\d*)%<\/p>", html)[1].replace(',', '.')) / 100, 5),
            round(float(regex_lookup_nogroup("<p>(\d*,\d*)%<\/p>", html)[2].replace(',', '.')) / 100, 5)
        ]

        # Heffingsvrij vermogen
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/vermogen_en_aanmerkelijk_belang/vermogen/belasting_betalen_over_uw_vermogen/heffingsvrij_vermogen/heffingsvrij_vermogen"
        html = read_html(url)

        self.heffingsvrijvermogen = regex_lookup_nogroup("<span>€ (.*?)<\/span>", html)[0]
        self.heffingsvrijvermogen = self.heffingsvrijvermogen * (self.belasting.fiscale_partner + 1)

        # CO2
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/auto_en_vervoer/belastingen_op_auto_en_motor/bpm/bpm_berekenen_en_betalen/bpm_tarief/bpm-tarief-personenauto"
        html = read_html(url)

        uitstoot_min = regex_lookup_nogroup("row\">\s*<p>(\d*)\s", html)
        uitstoot_bedragen = regex_lookup_nogroup(">€&nbsp;(.*?)<\/p>", html)
        self.CO2_BPM = [
            [uitstoot_min[0], uitstoot_bedragen[0], uitstoot_bedragen[1]],
            [uitstoot_min[1], uitstoot_bedragen[2], uitstoot_bedragen[3]],
            [uitstoot_min[2], uitstoot_bedragen[4], uitstoot_bedragen[5]],
            [uitstoot_min[3], uitstoot_bedragen[6], uitstoot_bedragen[7]],
            [uitstoot_min[4], uitstoot_bedragen[8], uitstoot_bedragen[9]]
        ]
        self.CO2_BPM = [[val.replace('.', '') for val in row] for row in self.CO2_BPM]  # replace dots
        self.CO2_BPM = [[int(val) for val in row] for row in self.CO2_BPM]  # convert to int
        self.CO2_diesel_grens = int(regex_lookup("(\d*)&nbsp;gram\/km\.<\/p>", html))
        self.CO2_diesel_toeslag = float(regex_lookup("van €&nbsp;(.*?)per gram", html).replace(',', '.'))

        # Algemene heffingskorting
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/inkomstenbelasting/heffingskortingen_boxen_tarieven/heffingskortingen/algemene_heffingskorting/tabel-algemene-heffingskorting-2019"
        html = read_html(url)

        alg_korting_values = regex_lookup_nogroup("€&nbsp;(.*?)<\/p>", html)
        alg_korting_values = [alg_korting_values[val].replace('.', '') for val in range(len(alg_korting_values))]
        self.alg_korting = [
            [alg_korting_values[0], alg_korting_values[2]],
            [alg_korting_values[1], float(regex_lookup_nogroup("<\/span>-(.*?)% x", html)[0].replace(',', '.')) / 100],
            [alg_korting_values[6], alg_korting_values[7]]
        ]
        self.alg_korting_aow = [
            [alg_korting_values[8], alg_korting_values[10]],
            [alg_korting_values[1], float(regex_lookup_nogroup("<\/span>-(.*?)% x", html)[1].replace(',', '.')) / 100],
            [alg_korting_values[14], alg_korting_values[15]]
        ]
        for row in range(len(self.alg_korting)):  # Convert all compatible numbers (not floats) to int
            for val in range(len(self.alg_korting[row])):
                if type(self.alg_korting[row][val]) == str:
                    self.alg_korting[row][val] = int(self.alg_korting[row][val])
                if type(self.alg_korting_aow[row][val]) == str:
                    self.alg_korting_aow[row][val] = int(self.alg_korting_aow[row][val])

        self.alg_korting[1][1] = self.alg_korting[0][1] - self.alg_korting[1][1] * (
                self.persoon.bruto_loon_jr - self.alg_korting[1][0])
        self.alg_korting_aow[1][1] = self.alg_korting_aow[0][1] - self.alg_korting_aow[1][1] * (
                self.persoon.bruto_loon_jr - self.alg_korting_aow[1][0])

        # Arbeidskorting
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/inkomstenbelasting/heffingskortingen_boxen_tarieven/heffingskortingen/arbeidskorting/tabel-arbeidskorting-2019"
        html = read_html(url)

        arbeidskorting_schaal_values = regex_lookup_nogroup("<p>€&nbsp;(?:<span>)?(.*?)(?:<\/span>)?<\/p>", html)
        self.arbeidskorting_schaal = [
            int(arbeidskorting_schaal_values[0].replace('.', '')),
            int(arbeidskorting_schaal_values[2].replace('.', '')),
            int(arbeidskorting_schaal_values[5].replace('.', '')),
            int(arbeidskorting_schaal_values[9].replace('.', ''))
        ]
        self.arbeidskorting_korting = [
            0.01754 * self.persoon.bruto_loon_jr,
            170 + 0.28712 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[1]),
            3399,
            3399 - 0.06 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[3]),
            0
        ]
        self.arbeidskorting_korting_aow = [
            0.00898 * self.persoon.bruto_loon_jr,
            88 + 0.14689 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[1]),
            1740,
            1740 - 0.03069 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[3]),
            0
        ]


class Calculation:

    def __init__(self, persoon, auto, belasting):
        self.persoon = persoon
        self.auto = auto
        self.belasting = belasting

    ''' 
        BTW wordt als volgt berekend:
        uitgaven met lage btw (voedsel etc) * 9%   +
        uitgaven met hogte btw (incidenteel, wasmachine etc) * 21%    +
        aantal gereden kilometers per jaar * verbruik (liter per km) * btw op benzine
    '''
    def get_BTW(self):
        self.btw_laag = self.persoon.uitgaven_laag * 0.09
        self.btw_hoog = self.persoon.uitgaven_hoog * 0.21 + self.auto.km_jaar * self.auto.verbruik * self.belasting.benzine_btw
        self.btw = self.btw_hoog + self.btw_laag
        print('Calculated BTW')

    # Roken / Alcohol
    # self.rokenalcohol = 52*(self.persoon.roken*self.belasting.roken_accijns*6.50 + self.persoon.alcohol*self.belasting.alcohol_accijns*0.3)*1.21
    print('Calculated roken/alcohol')

    ''' 
        Brandstof accijns is verbruik * kilometers * accijns
        BPM wordt berekend per CO2 uitstoot met verschillende schalen
        Als je een diesel rijd die meer dan 61 g/km CO2 uitstoot betaal je extra dieseltoeslag
    '''
    def get_auto(self):
        if self.auto.brandstof.lower() == 'benzine':
            self.brandstof_accijns = self.auto.km_jaar * self.auto.verbruik * self.belasting.benzine_accijns
        elif self.auto.brandstof.lower() == 'diesel':
            self.brandstof_accijns = self.auto.km_jaar * self.auto.verbruik * self.belasting.diesel_accijns

        if self.auto.CO2_uitstoot == 0:
            self.BPM = 0
        else:
            row = find_row(self.belasting.CO2_BPM, self.auto.CO2_uitstoot)
            self.BPM = (self.auto.CO2_uitstoot - self.belasting.CO2_BPM[row][0]) * self.belasting.CO2_BPM[row][2] + \
                       self.belasting.CO2_BPM[row][1]

        if self.auto.CO2_uitstoot > self.belasting.CO2_diesel_grens and self.auto.brandstof.lower() == 'diesel':
            self.BPM = self.BPM + self.belasting.CO2_diesel_toeslag * (
                    self.auto.CO2_uitstoot - self.belasting.CO2_diesel_grens)
        print('Calculated car')

    # Loon
    def get_loon(self):
        if self.persoon.leeftijd > 67:
            procent = self.belasting.loonbelasting_aow
            alg_korting = self.belasting.alg_korting_aow
            arbeid = self.belasting.arbeidskorting_korting_aow
        else:
            procent = self.belasting.loonbelasting
            alg_korting = self.belasting.alg_korting
            arbeid = self.belasting.arbeidskorting_korting

        loon_bruto = self.persoon.bruto_loon_jr
        loon_totaal = self.persoon.loon_totaal
        loon_speciaal = self.persoon.loon_speciaal
        schaal_loon = self.belasting.loonbelasting_schaal
        schaal_arbeid = self.belasting.arbeidskorting_schaal

        row = find_row(schaal_loon, loon_bruto)
        over = loon_bruto - schaal_loon[row]
        loontaks = 0
        if row == 0:
            loontaks += (schaal_loon[row + 1] - schaal_loon[row]) * procent[row]
        else:
            for schijf in range(row):
                loontaks += (schaal_loon[schijf + 1] - schaal_loon[schijf]) * procent(schijf)
        loontaks += over * procent[row]

        row = find_row(schaal_loon, loon_totaal)
        loontaks_speciaal = procent[row] * loon_speciaal

        row = find_row(alg_korting, loon_bruto)
        heffingskorting = alg_korting[row][1]

        row = find_row(schaal_arbeid, arbeid)
        arbeidskorting = arbeid[row]

        # Vermogensbelasting
        vermogen = self.persoon.spaargeld - self.persoon.schulden - self.belasting.heffingsvrijvermogen
        vermogensbelasting = 0
        if vermogen > 0:
            row = find_row(self.belasting.vermogensbelasting_schaal, vermogen)
            over = vermogen - self.belasting.vermogensbelasting_schaal[row]
            if row == 0:
                vermogensbelasting += self.belasting.vermogensbelasting_schaal[row] * \
                                      self.belasting.vermogensbelasting_rendement[row]
            else:
                for schijf in range(row):
                    vermogensbelasting += (self.belasting.vermogensbelasting_schaal[schijf + 1] -
                                           self.belasting.vermogensbelasting_schaal[schijf]) * \
                                          self.belasting.vermogensbelasting_rendement[schijf]
                vermogensbelasting += over * self.belasting.vermogensbelasting_rendement[row]


def regex_lookup(regex_string, data_to_search):
    reg = re.compile(r"" + regex_string + "", re.DOTALL)
    data = reg.search(data_to_search)
    return data.group(1)


def regex_lookup_nogroup(regex_string, data_to_search):
    reg = re.compile(r"" + regex_string + "", re.DOTALL)
    data = reg.findall(data_to_search)
    return data


def read_html(url):
    response = urllib.request.urlopen(url)
    html = response.read().decode(response.headers.get_content_charset())
    return html


def find_row(table, var):
    for row in range(len(table) - 1, -1, -1):  # count down from rows in table to 0
        try:
            if var > table[row][0]:
                return row
        except TypeError:
            if var > table[row]:
                return row


def check_input(arguments, str_idx, int_idx):
    arg_list_keys = list(arguments.keys())[1:]  # don't include self
    arg_list_values = list(arguments.values())[1:]
    for idx in str_idx:
        while True:
            if type(arg_list_values[idx]) == str:
                break
            else:
                arg_list_values[idx] = input(
                    'Value \'{}\' of variable \'{}\' is not the correct type. Please insert letters only >>> '.format(
                        arg_list_values[idx], arg_list_keys[idx]))
                arguments[arg_list_keys[idx]] = arg_list_values[idx]
                continue
    print('Checked all strings\n')
    for idx in int_idx:
        while True:
            if type(arg_list_values[idx]) == int:
                break
            else:
                arg_list_values[idx] = input(
                    'Value \'{}\' of variable \'{}\' is not the correct type. Please insert numbers only >>> '.format(
                        arg_list_values[idx], arg_list_keys[idx]))
                try:
                    arg_list_values[idx] = int(arg_list_values[idx])
                    arguments[arg_list_keys[idx]] = arg_list_values[idx]
                    break
                except ValueError:
                    continue
    print('Checked all ints\n')
    return arguments
